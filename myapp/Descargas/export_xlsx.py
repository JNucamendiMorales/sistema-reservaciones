import io
import math
from datetime import datetime, timedelta
import pytz

from django.utils import timezone
from django.db.models import Count, Sum

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.styles import Font, PatternFill

from myapp.models import Salon, Reservacion
from django.contrib.auth.models import User


def _nice_step(max_val, prefer_two=False):
    if prefer_two:
        return 2
    if not max_val or max_val <= 0:
        return 1
    approx = max_val / 10.0
    magnitude = 10 ** math.floor(math.log10(approx))
    residual = approx / magnitude
    if residual <= 1:
        nice = 1
    elif residual <= 2:
        nice = 2
    elif residual <= 5:
        nice = 5
    else:
        nice = 10
    return int(nice * magnitude)


def exportar_xlsx_nativo(charts_data, fecha_descarga=None):
    wb = Workbook()
    ws_default = wb.active
    wb.remove(ws_default)

    for chart in charts_data:
        ws = wb.create_sheet(title=chart["nombre"][:31])
        ws.append(["Etiqueta", "Valor"])

        for fila in chart["datos"]:
            ws.append([fila.get("Etiqueta", ""), fila.get("Valor", "")])

        name = chart["nombre"].lower()
        if "line" in name or "línea" in name:
            c = LineChart()
        elif "pie" in name or "pastel" in name:
            c = PieChart()
        else:
            c = BarChart()

        c.title = chart["nombre"]

        try:
            if isinstance(c, (BarChart, LineChart)):
                c.x_axis.tickLblPos = "low"
        except Exception:
            pass

        try:
            c.legend = None
        except Exception:
            pass

        if isinstance(c, PieChart):
            try:
                c.width = 8
                c.height = 6
            except Exception:
                pass

        data = Reference(ws, min_col=2, min_row=2, max_row=len(chart["datos"]) + 1)
        categories = Reference(ws, min_col=1, min_row=2, max_row=len(chart["datos"]) + 1)

        c.add_data(data, titles_from_data=False)
        c.set_categories(categories)

        ws.add_chart(c, "D2")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    from django.http import HttpResponse
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="graficas.xlsx"'
    return response


def exportar_xlsx_reservaciones_semanal(fecha_descarga=None):
    tz = pytz.timezone('America/Mexico_City')
    hoy_local = timezone.now().astimezone(tz).date()
    inicio_semana = hoy_local - timedelta(days=hoy_local.weekday())
    dias_nombres = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]

    conteos = []
    for i in range(7):
        d = inicio_semana + timedelta(days=i)
        total = Reservacion.objects.filter(fecha_reserva=d).count()
        conteos.append(int(total))

    por_salon_qs = Reservacion.objects.filter(
        fecha_reserva__range=[inicio_semana, inicio_semana + timedelta(days=6)]
    ).values("salon__nombre").annotate(total=Count("id")).order_by("-total")[:10]
    datos_salones = list(por_salon_qs)

    ingresos = []
    for i in range(7):
        d = inicio_semana + timedelta(days=i)
        total_ing = Reservacion.objects.filter(fecha_reserva=d).aggregate(total=Sum("precio_total"))["total"] or 0.0
        ingresos.append(float(total_ing))

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Reservaciones"
    if fecha_descarga:
        ws1["F1"] = str(fecha_descarga)
    ws1.append(["Reservaciones por periodo (Semanal)"])
    ws1.append([])

    # TABLA 1: Encabezados con color azul
    ws1['A4'] = "Dia"
    ws1['B4'] = "Total"
    ws1['A4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws1['B4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws1['A4'].font = Font(color="FFFFFF", bold=True)
    ws1['B4'].font = Font(color="FFFFFF", bold=True)
    
    for i, dia in enumerate(dias_nombres):
        ws1[f'A{5+i}'] = dia
        ws1[f'B{5+i}'] = conteos[i]
    total_res = sum(conteos)
    ws1.append([])
    ws1.append(["TOTAL", total_res])

    chart1 = BarChart()
    chart1.title = "Reservaciones - Semana"
    chart1.y_axis.title = None
    chart1.x_axis.title = None

    try:
        max_count = max(conteos) if conteos else 0
        y_max = int(max_count)
        if y_max % 2 != 0:
            y_max += 1
        if y_max == 0:
            y_max = 2
        chart1.y_axis.scaling.minVal = 0
        chart1.y_axis.scaling.maxVal = y_max
        chart1.y_axis.majorUnit = 2
        chart1.y_axis.tickLblPos = "low"
        chart1.y_axis.delete = False
        try:
            chart1.y_axis.crosses = "min"
        except Exception:
            pass
        chart1.y_axis.majorTickMark = "out"
        chart1.y_axis.majorGridlines.graphicalProperties.ln.solidFill = "D3D3D3"
        chart1.y_axis.majorGridlines.graphicalProperties.ln.w = 12700
    except Exception:
        pass

    try:
        chart1.x_axis.tickLblPos = "low"
        chart1.x_axis.majorTickMark = "out"
        chart1.x_axis.delete = False
    except Exception:
        pass

    try:
        chart1.legend = None
    except Exception:
        pass

    data_ref = Reference(ws1, min_col=2, min_row=5, max_row=5 + len(conteos) - 1)
    cats_ref = Reference(ws1, min_col=1, min_row=5, max_row=5 + len(conteos) - 1)
    chart1.add_data(data_ref, titles_from_data=False)
    chart1.set_categories(cats_ref)

    try:
        chart1.gapWidth = 50
    except Exception:
        pass
    chart1.height = 10
    chart1.width = 18
    ws1.add_chart(chart1, "D4")

    ws2 = wb.create_sheet("Salones")
    if fecha_descarga:
        ws2["F1"] = str(fecha_descarga)
    ws2.append(["Salones más reservados (Semanal)"])
    ws2.append([])
    
    # TABLA 2: Encabezados con color azul
    ws2['A4'] = "Salón"
    ws2['B4'] = "Reservaciones"
    ws2['A4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws2['B4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws2['A4'].font = Font(color="FFFFFF", bold=True)
    ws2['B4'].font = Font(color="FFFFFF", bold=True)
    
    # Escribir datos reales de salones EMPEZANDO EN FILA 5
    row_start = 5
    for idx, s in enumerate(datos_salones):
        ws2[f'A{row_start + idx}'] = s["salon__nombre"]
        ws2[f'B{row_start + idx}'] = int(s["total"])
    
    total_sal = sum(int(s["total"]) for s in datos_salones) if datos_salones else 0
    ws2.append([])
    ws2.append(["TOTAL", total_sal])

    # PIE CHART - REFERENCIAS CORRECTAS
    pie = PieChart()
    pie.title = "Salones más reservados"
    try:
        pie.width = 18
        pie.height = 10
        pie.legend.position = "r"
        pie.legend.overlay = False
        pie.layout = Layout(manualLayout=ManualLayout())
        pie.layout.manualLayout.x = 0.02
        pie.layout.manualLayout.y = 0.05
        pie.layout.manualLayout.w = 0.60
        pie.layout.manualLayout.h = 0.90
    except Exception:
        pass

    if datos_salones:
        # Datos comienzan en B5, terminan en B(5 + len(datos_salones) - 1)
        data_ref = Reference(ws2, min_col=2, min_row=row_start, max_row=row_start + len(datos_salones) - 1)
        # Categorías: nombres de salones en A5:A(5 + len(datos_salones) - 1)
        cats_ref = Reference(ws2, min_col=1, min_row=row_start, max_row=row_start + len(datos_salones) - 1)
        pie.add_data(data_ref, titles_from_data=False)
        pie.set_categories(cats_ref)
    ws2.add_chart(pie, "D4")

    ws3 = wb.create_sheet("Ingresos")
    if fecha_descarga:
        ws3["F1"] = str(fecha_descarga)
    ws3.append(["Ingresos semanalmente"])
    ws3.append([])
    
    # TABLA 3: Encabezados con color azul
    ws3['A4'] = "Día"
    ws3['B4'] = "Ingresos"
    ws3['A4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws3['B4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws3['A4'].font = Font(color="FFFFFF", bold=True)
    ws3['B4'].font = Font(color="FFFFFF", bold=True)
    
    for nombre, valor in zip(dias_nombres, ingresos):
        ws3.append([nombre, float(valor)])
    total_ing = sum(ingresos)
    ws3.append([])
    ws3.append(["TOTAL", float(total_ing)])

    line = LineChart()
    line.title = "Ingresos - Semana"
    line.y_axis.title = None
    line.x_axis.title = None

    try:
        max_ing = max(ingresos) if ingresos else 0
        step = _nice_step(max_ing, prefer_two=False)
        y_max_ing = int(((int(max_ing) + step - 1) // step) * step) if step else int(max_ing)
        if y_max_ing == 0:
            y_max_ing = step or 1
        line.y_axis.scaling.minVal = 0
        line.y_axis.scaling.maxVal = y_max_ing
        line.y_axis.majorUnit = step
        line.y_axis.tickLblPos = "low"
        line.y_axis.delete = False
        try:
            line.y_axis.crosses = "min"
        except Exception:
            pass
        line.x_axis.tickLblPos = "low"
        line.x_axis.majorTickMark = "out"
        line.y_axis.majorTickMark = "out"
        line.y_axis.majorGridlines.graphicalProperties.ln.solidFill = "D3D3D3"
        line.y_axis.majorGridlines.graphicalProperties.ln.w = 12700
    except Exception:
        pass

    try:
        line.x_axis.tickLblPos = "low"
        line.x_axis.majorTickMark = "out"
        line.x_axis.delete = False
    except Exception:
        pass

    try:
        line.legend = None
    except Exception:
        pass

    data_ref = Reference(ws3, min_col=2, min_row=4, max_row=4 + len(ingresos) - 1)
    cats_ref = Reference(ws3, min_col=1, min_row=4, max_row=4 + len(ingresos) - 1)
    line.add_data(data_ref, titles_from_data=False)
    line.set_categories(cats_ref)

    try:
        for s in line.series:
            s.smooth = False
            try:
                s.graphicalProperties.line.width = 20000
            except Exception:
                pass
    except Exception:
        pass

    line.height = 10
    line.width = 18
    ws3.add_chart(line, "D4")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    from django.http import HttpResponse
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="reservaciones_semanal.xlsx"'
    return response


def exportar_xlsx_reservaciones_mensual(fecha_descarga=None):
    tz = pytz.timezone('America/Mexico_City')
    fecha_actual = timezone.now().astimezone(tz)
    mes_actual = fecha_actual.month
    año_actual = fecha_actual.year

    dias_rango = [
        (1, 5, "1-5"),
        (6, 10, "6-10"),
        (11, 15, "11-15"),
        (16, 20, "16-20"),
        (21, 25, "21-25"),
        (26, 31, "26-31"),
    ]

    conteos = []
    for start, end, label in dias_rango:
        total = Reservacion.objects.filter(
            fecha_reserva__year=año_actual,
            fecha_reserva__month=mes_actual,
            fecha_reserva__day__gte=start,
            fecha_reserva__day__lte=end
        ).count()
        conteos.append(int(total))

    por_salon_qs = Reservacion.objects.filter(
        fecha_reserva__year=año_actual,
        fecha_reserva__month=mes_actual
    ).values("salon__nombre").annotate(total=Count("id")).order_by("-total")[:10]
    datos_salones = list(por_salon_qs)

    ingresos = []
    for start, end, label in dias_rango:
        total = Reservacion.objects.filter(
            fecha_reserva__year=año_actual,
            fecha_reserva__month=mes_actual,
            fecha_reserva__day__gte=start,
            fecha_reserva__day__lte=end
        ).aggregate(total=Sum("precio_total"))["total"] or 0.0
        ingresos.append(float(total))

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Reservaciones"
    if fecha_descarga:
        ws1["F1"] = str(fecha_descarga)
    ws1.append([f"Reservaciones por periodo - {fecha_actual.strftime('%B %Y')}"])
    ws1.append([])
    
    # TABLA 1: Encabezados con color azul
    ws1['A4'] = "Dias"
    ws1['B4'] = "Reservaciones"
    ws1['A4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws1['B4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws1['A4'].font = Font(color="FFFFFF", bold=True)
    ws1['B4'].font = Font(color="FFFFFF", bold=True)
    
    for label, valor in zip([r[2] for r in dias_rango], conteos):
        ws1.append([label, valor])
    total_res = sum(conteos)
    ws1.append([])
    ws1.append(["TOTAL", total_res])

    chart = BarChart()
    chart.type = "col"
    chart.style = 2
    chart.title = "Reservaciones por periodo (Mensual)"
    chart.y_axis.title = None
    chart.x_axis.title = None

    try:
        max_count_m = max(conteos) if conteos else 0
        y_max_m = int(max_count_m)
        if y_max_m % 2 != 0:
            y_max_m += 1
        if y_max_m == 0:
            y_max_m = 2
        chart.y_axis.scaling.minVal = 0
        chart.y_axis.scaling.maxVal = y_max_m
        chart.y_axis.majorUnit = 2
        chart.y_axis.tickLblPos = "low"
        chart.y_axis.delete = False
        try:
            chart.y_axis.crosses = "min"
        except Exception:
            pass
        chart.x_axis.tickLblPos = "low"
        try:
            chart.x_axis.delete = False
        except Exception:
            pass
        chart.x_axis.majorTickMark = "out"
        chart.y_axis.majorTickMark = "out"
        chart.y_axis.majorGridlines.graphicalProperties.ln.solidFill = "D3D3D3"
        chart.y_axis.majorGridlines.graphicalProperties.ln.w = 12700
        try:
            chart.x_axis.crosses = "max"
        except Exception:
            pass
    except Exception:
        pass
    try:
        chart.legend = None
    except Exception:
        pass

    data_ref = Reference(ws1, min_col=2, min_row=5, max_row=5 + len(conteos) - 1)
    categories_ref = Reference(ws1, min_col=1, min_row=5, max_row=5 + len(conteos) - 1)

    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(categories_ref)

    try:
        chart.gapWidth = 50
    except Exception:
        pass
    chart.height = 10
    chart.width = 18

    ws1.add_chart(chart, "D6")

    ws2 = wb.create_sheet("Salones")
    if fecha_descarga:
        ws2["F1"] = str(fecha_descarga)
    ws2.append(["Salones más reservados mensualmente"])
    ws2.append([])
    
    # TABLA 2: Encabezados con color azul
    ws2['A4'] = "Salón"
    ws2['B4'] = "Reservaciones"
    ws2['A4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws2['B4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws2['A4'].font = Font(color="FFFFFF", bold=True)
    ws2['B4'].font = Font(color="FFFFFF", bold=True)
    
    total_salones = 0
    for s in datos_salones:
        ws2.append([s["salon__nombre"], s["total"]])
        total_salones += s["total"]
    ws2.append([])
    ws2.append(["TOTAL", total_salones])

    try:
        pie = PieChart()
        pie.title = "Salones más reservados"
        pie.width = 18
        pie.height = 10
        pie.legend.position = "r"
        pie.legend.overlay = False
        pie.layout = Layout(manualLayout=ManualLayout())
        pie.layout.manualLayout.x = 0.02
        pie.layout.manualLayout.y = 0.05
        pie.layout.manualLayout.w = 0.60
        pie.layout.manualLayout.h = 0.90

        if datos_salones:
            start_row = 4
            end_row = start_row + len(datos_salones) - 1
            data_ref = Reference(ws2, min_col=2, min_row=start_row, max_row=end_row)
            cats_ref = Reference(ws2, min_col=1, min_row=start_row, max_row=end_row)
            pie.add_data(data_ref, titles_from_data=False)
            pie.set_categories(cats_ref)

        ws2.add_chart(pie, "D4")
    except Exception:
        pass

    ws3 = wb.create_sheet("Ingresos")
    if fecha_descarga:
        ws3["F1"] = str(fecha_descarga)
    ws3.append(["Ingresos por Usuario (Mensual)"])
    ws3.append([])
    
    # TABLA 3: Encabezados con color azul
    ws3['A4'] = "Usuario"
    ws3['B4'] = "Ingresos"
    ws3['A4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws3['B4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws3['A4'].font = Font(color="FFFFFF", bold=True)
    ws3['B4'].font = Font(color="FFFFFF", bold=True)
    
    for usuario in datos_ingresos:
        ws3.append([usuario["usuario__username"], float(usuario["total"] or 0.0)])
    total_ing = sum(float(u["total"] or 0.0) for u in datos_ingresos)
    ws3.append([])
    ws3.append(["TOTAL", float(total_ing)])

    line = LineChart()
    line.title = "Ingresos por Usuario"
    line.y_axis.title = None
    line.x_axis.title = None

    try:
        ingresos_vals = [float(u["total"] or 0.0) for u in datos_ingresos]
        max_ing_m = max(ingresos_vals) if ingresos_vals else 0
        step_m = _nice_step(max_ing_m, prefer_two=False)
        y_max_ing_m = int(((int(max_ing_m) + step_m - 1) // step_m) * step_m) if step_m else int(max_ing_m)
        if y_max_ing_m == 0:
            y_max_ing_m = step_m or 1
        line.y_axis.scaling.minVal = 0
        line.y_axis.scaling.maxVal = y_max_ing_m
        line.y_axis.majorUnit = step_m
        line.y_axis.tickLblPos = "low"
        line.y_axis.delete = False
        try:
            line.y_axis.crosses = "min"
        except Exception:
            pass
        line.x_axis.tickLblPos = "low"
        line.x_axis.majorTickMark = "out"
        line.y_axis.majorTickMark = "out"
        line.y_axis.majorGridlines.graphicalProperties.ln.solidFill = "D3D3D3"
        line.y_axis.majorGridlines.graphicalProperties.ln.w = 12700
    except Exception:
        pass

    try:
        line.x_axis.tickLblPos = "low"
        line.x_axis.majorTickMark = "out"
        line.x_axis.delete = False
    except Exception:
        pass

    try:
        line.legend = None
    except Exception:
        pass

    data_ref = Reference(ws3, min_col=2, min_row=4, max_row=4 + len(datos_ingresos) - 1)
    cats_ref = Reference(ws3, min_col=1, min_row=4, max_row=4 + len(datos_ingresos) - 1)
    line.add_data(data_ref, titles_from_data=False)
    line.set_categories(cats_ref)

    try:
        for s in line.series:
            s.smooth = False
            try:
                s.graphicalProperties.line.width = 20000
            except Exception:
                pass
    except Exception:
        pass

    line.height = 10
    line.width = 18
    ws3.add_chart(line, "D4")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    from django.http import HttpResponse
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="reservaciones_mensual.xlsx"'
    return response


def exportar_xlsx_salones_semanal(fecha_descarga=None):
    """Salones semanal con gráficos - estructura igual a reservaciones/usuarios"""
    from datetime import timedelta
    from django.utils import timezone
    from django.db.models import Count, Sum, Avg
    import pytz
    import io
    
    tz = pytz.timezone('America/Mexico_City')
    hoy = timezone.now().astimezone(tz).date()
    inicio_semana = hoy - timedelta(days=hoy.weekday())
    dias_nombres = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]

    # Datos: Top 5 calificación
    top_calificacion = list(
        Salon.objects.values("nombre", "calificacion").order_by("-calificacion")[:5]
    )

    # Datos: Top 5 precios
    top_precios = list(
        Salon.objects.values("nombre", "precio").order_by("-precio")[:5]
    )

    # Datos: Top 5 ingresos (semanal)
    top_ingresos = list(
        Reservacion.objects.filter(
            fecha_reserva__range=[inicio_semana, inicio_semana + timedelta(days=6)]
        ).values("salon__nombre").annotate(total=Sum("precio_total")).order_by("-total")[:5]
    )

    wb = Workbook()
    
    # HOJA 1: Calificación
    ws1 = wb.active
    ws1.title = "Calificación"
    if fecha_descarga:
        ws1["F1"] = str(fecha_descarga)
    ws1['A1'] = "Salones con mejores reseñas (Semanal)"
    ws1['A1'].font = Font(bold=True, size=12)
    
    ws1['A4'] = "Salón"
    ws1['B4'] = "Calificación"
    ws1['A4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws1['B4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws1['A4'].font = Font(color="FFFFFF", bold=True)
    ws1['B4'].font = Font(color="FFFFFF", bold=True)
    
    row_cal = 5
    total_cal = 0.0
    for salon in top_calificacion:
        ws1[f'A{row_cal}'] = salon["nombre"]
        cal_val = float(salon["calificacion"] or 0.0)
        ws1[f'B{row_cal}'] = round(cal_val, 2)
        total_cal += cal_val
        row_cal += 1

    # REMOVED: total row for Calificación (no tiene sentido sumar calificaciones)
    # previously: ws1[f'A{row_cal}'] = "TOTAL" / ws1[f'B{row_cal}'] = round(total_cal,2)

    chart1 = BarChart()
    chart1.title = "Salones - Calificación"
    chart1.y_axis.title = None
    chart1.x_axis.title = None

    try:
        max_cal = max([float(s["calificacion"] or 0.0) for s in top_calificacion]) if top_calificacion else 0
        y_max_cal = int(max_cal) + 1 if max_cal > 0 else 5
        chart1.y_axis.scaling.minVal = 0
        chart1.y_axis.scaling.maxVal = y_max_cal
        chart1.y_axis.majorUnit = 1
        chart1.y_axis.tickLblPos = "low"
        chart1.y_axis.delete = False
        chart1.y_axis.majorTickMark = "out"
        chart1.y_axis.majorGridlines.graphicalProperties.ln.solidFill = "D3D3D3"
        chart1.y_axis.majorGridlines.graphicalProperties.ln.w = 12700
    except Exception:
        pass

    try:
        chart1.x_axis.tickLblPos = "low"
        chart1.x_axis.majorTickMark = "out"
        chart1.x_axis.delete = False
        chart1.legend = None
    except Exception:
        pass

    data_ref = Reference(ws1, min_col=2, min_row=5, max_row=row_cal - 1)
    cats_ref = Reference(ws1, min_col=1, min_row=5, max_row=row_cal - 1)
    chart1.add_data(data_ref, titles_from_data=False)
    chart1.set_categories(cats_ref)

    try:
        chart1.gapWidth = 50
    except Exception:
        pass
    chart1.height = 10
    chart1.width = 18
    ws1.add_chart(chart1, "D4")

    # HOJA 2: Precios
    ws2 = wb.create_sheet("Precios")
    if fecha_descarga:
        ws2["F1"] = str(fecha_descarga)
    ws2['A1'] = "Salones más caros (Semanal)"
    ws2['A1'].font = Font(bold=True, size=12)
    
    ws2['A4'] = "Salón"
    ws2['B4'] = "Precio (MXN)"
    ws2['A4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws2['B4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws2['A4'].font = Font(color="FFFFFF", bold=True)
    ws2['B4'].font = Font(color="FFFFFF", bold=True)
    
    row_pre = 5
    total_pre = 0.0
    for salon in top_precios:
        ws2[f'A{row_pre}'] = salon["nombre"]
        pre_val = float(salon["precio"] or 0.0)
        ws2[f'B{row_pre}'] = pre_val
        total_pre += pre_val
        row_pre += 1
    
    ws2[f'A{row_pre}'] = "TOTAL"
    ws2[f'A{row_pre}'].font = Font(bold=True)
    ws2[f'B{row_pre}'] = total_pre
    ws2[f'B{row_pre}'].font = Font(bold=True)

    chart2 = BarChart()
    chart2.title = "Salones - Precios"
    chart2.y_axis.title = None
    chart2.x_axis.title = None

    try:
        max_pre = max([float(s["precio"] or 0.0) for s in top_precios]) if top_precios else 0
        step_pre = _nice_step(max_pre, prefer_two=False)
        y_max_pre = int(((int(max_pre) + step_pre - 1) // step_pre) * step_pre) if step_pre else int(max_pre)
        if y_max_pre == 0:
            y_max_pre = step_pre or 1000
        chart2.y_axis.scaling.minVal = 0
        chart2.y_axis.scaling.maxVal = y_max_pre
        chart2.y_axis.majorUnit = step_pre
        chart2.y_axis.tickLblPos = "low"
        chart2.y_axis.delete = False
        chart2.y_axis.majorTickMark = "out"
        chart2.y_axis.majorGridlines.graphicalProperties.ln.solidFill = "D3D3D3"
        chart2.y_axis.majorGridlines.graphicalProperties.ln.w = 12700
    except Exception:
        pass

    try:
        chart2.x_axis.tickLblPos = "low"
        chart2.x_axis.majorTickMark = "out"
        chart2.x_axis.delete = False
        chart2.legend = None
    except Exception:
        pass

    data_ref = Reference(ws2, min_col=2, min_row=5, max_row=row_pre - 1)
    cats_ref = Reference(ws2, min_col=1, min_row=5, max_row=row_pre - 1)
    chart2.add_data(data_ref, titles_from_data=False)
    chart2.set_categories(cats_ref)

    try:
        chart2.gapWidth = 50
    except Exception:
        pass
    chart2.height = 10
    chart2.width = 18
    ws2.add_chart(chart2, "D4")

    # HOJA 3: Ingresos
    ws3 = wb.create_sheet("Ingresos")
    if fecha_descarga:
        ws3["F1"] = str(fecha_descarga)
    ws3['A1'] = "Salones con más Ingresos (Semanal)"
    ws3['A1'].font = Font(bold=True, size=12)
    
    ws3['A4'] = "Salón"
    ws3['B4'] = "Ingresos (MXN)"
    ws3['A4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws3['B4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws3['A4'].font = Font(color="FFFFFF", bold=True)
    ws3['B4'].font = Font(color="FFFFFF", bold=True)
    
    row_ing = 5
    total_ing = 0.0
    for salon in top_ingresos:
        ws3[f'A{row_ing}'] = salon["salon__nombre"]
        ing_val = float(salon["total"] or 0.0)
        ws3[f'B{row_ing}'] = ing_val
        total_ing += ing_val
        row_ing += 1
    
    ws3[f'A{row_ing}'] = "TOTAL"
    ws3[f'A{row_ing}'].font = Font(bold=True)
    ws3[f'B{row_ing}'] = total_ing
    ws3[f'B{row_ing}'].font = Font(bold=True)

    chart3 = BarChart()
    chart3.title = "Salones - Ingresos Semanal"
    chart3.y_axis.title = None
    chart3.x_axis.title = None

    try:
        ing_vals = [float(s["total"] or 0.0) for s in top_ingresos]
        max_ing = max(ing_vals) if ing_vals else 0
        step_ing = _nice_step(max_ing, prefer_two=False)
        y_max_ing = int(((int(max_ing) + step_ing - 1) // step_ing) * step_ing) if step_ing else int(max_ing)
        if y_max_ing == 0:
            y_max_ing = step_ing or 1000
        chart3.y_axis.scaling.minVal = 0
        chart3.y_axis.scaling.maxVal = y_max_ing
        chart3.y_axis.majorUnit = step_ing
        chart3.y_axis.tickLblPos = "low"
        chart3.y_axis.delete = False
        chart3.y_axis.majorTickMark = "out"
        chart3.y_axis.majorGridlines.graphicalProperties.ln.solidFill = "D3D3D3"
        chart3.y_axis.majorGridlines.graphicalProperties.ln.w = 12700
    except Exception:
        pass

    try:
        chart3.x_axis.tickLblPos = "low"
        chart3.x_axis.majorTickMark = "out"
        chart3.x_axis.delete = False
        chart3.legend = None
    except Exception:
        pass

    data_ref = Reference(ws3, min_col=2, min_row=5, max_row=row_ing - 1)
    cats_ref = Reference(ws3, min_col=1, min_row=5, max_row=row_ing - 1)
    chart3.add_data(data_ref, titles_from_data=False)
    chart3.set_categories(cats_ref)

    try:
        chart3.gapWidth = 50
    except Exception:
        pass
    chart3.height = 10
    chart3.width = 18
    ws3.add_chart(chart3, "D4")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    from django.http import HttpResponse
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="salones_semanal.xlsx"'
    return response


def exportar_xlsx_salones_mensual(fecha_descarga=None):
    """Salones mensual con gráficos - estructura igual a reservaciones/usuarios"""
    from django.db.models import Count, Sum, Avg
    from django.utils import timezone
    import pytz
    import io
    
    tz = pytz.timezone('America/Mexico_City')
    fecha_actual = timezone.now().astimezone(tz)
    mes_actual = fecha_actual.month
    año_actual = fecha_actual.year

    # Datos: Top 5 calificación
    top_calificacion = list(
        Salon.objects.values("nombre", "calificacion").order_by("-calificacion")[:5]
    )

    # Datos: Top 5 precios
    top_precios = list(
        Salon.objects.values("nombre", "precio").order_by("-precio")[:5]
    )

    # Datos: Top 5 ingresos (mensual)
    top_ingresos = list(
        Reservacion.objects.filter(
            fecha_reserva__year=año_actual,
            fecha_reserva__month=mes_actual
        ).values("salon__nombre").annotate(total=Sum("precio_total")).order_by("-total")[:5]
    )

    wb = Workbook()
    
    # HOJA 1: Calificación
    ws1 = wb.active
    ws1.title = "Calificación"
    if fecha_descarga:
        ws1["F1"] = str(fecha_descarga)
    ws1['A1'] = f"Salones con mejores reseñas - {fecha_actual.strftime('%B %Y')}"
    ws1['A1'].font = Font(bold=True, size=12)
    
    ws1['A4'] = "Salón"
    ws1['B4'] = "Calificación"
    ws1['A4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws1['B4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws1['A4'].font = Font(color="FFFFFF", bold=True)
    ws1['B4'].font = Font(color="FFFFFF", bold=True)
    
    row_cal = 5
    total_cal = 0.0
    for salon in top_calificacion:
        ws1[f'A{row_cal}'] = salon["nombre"]
        cal_val = float(salon["calificacion"] or 0.0)
        ws1[f'B{row_cal}'] = round(cal_val, 2)
        total_cal += cal_val
        row_cal += 1

    # REMOVED: total row for Calificación (no tiene sentido sumar calificaciones)
    # previously: ws1[f'A{row_cal}'] = "TOTAL" / ws1[f'B{row_cal}'] = round(total_cal,2)

    chart1 = BarChart()
    chart1.title = "Salones - Calificación (Mensual)"
    chart1.y_axis.title = None
    chart1.x_axis.title = None

    try:
        max_cal = max([float(s["calificacion"] or 0.0) for s in top_calificacion]) if top_calificacion else 0
        y_max_cal = int(max_cal) + 1 if max_cal > 0 else 5
        chart1.y_axis.scaling.minVal = 0
        chart1.y_axis.scaling.maxVal = y_max_cal
        chart1.y_axis.majorUnit = 1
        chart1.y_axis.tickLblPos = "low"
        chart1.y_axis.delete = False
        chart1.y_axis.majorTickMark = "out"
        chart1.y_axis.majorGridlines.graphicalProperties.ln.solidFill = "D3D3D3"
        chart1.y_axis.majorGridlines.graphicalProperties.ln.w = 12700
    except Exception:
        pass

    try:
        chart1.x_axis.tickLblPos = "low"
        chart1.x_axis.majorTickMark = "out"
        chart1.x_axis.delete = False
        chart1.legend = None
    except Exception:
        pass

    data_ref = Reference(ws1, min_col=2, min_row=5, max_row=row_cal - 1)
    cats_ref = Reference(ws1, min_col=1, min_row=5, max_row=row_cal - 1)
    chart1.add_data(data_ref, titles_from_data=False)
    chart1.set_categories(cats_ref)

    try:
        chart1.gapWidth = 50
    except Exception:
        pass
    chart1.height = 10
    chart1.width = 18
    ws1.add_chart(chart1, "D4")

    # HOJA 2: Precios
    ws2 = wb.create_sheet("Precios")
    if fecha_descarga:
        ws2["F1"] = str(fecha_descarga)
    ws2['A1'] = f"Salones más caros - {fecha_actual.strftime('%B %Y')}"
    ws2['A1'].font = Font(bold=True, size=12)
    
    ws2['A4'] = "Salón"
    ws2['B4'] = "Precio (MXN)"
    ws2['A4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws2['B4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws2['A4'].font = Font(color="FFFFFF", bold=True)
    ws2['B4'].font = Font(color="FFFFFF", bold=True)
    
    row_pre = 5
    total_pre = 0.0
    for salon in top_precios:
        ws2[f'A{row_pre}'] = salon["nombre"]
        pre_val = float(salon["precio"] or 0.0)
        ws2[f'B{row_pre}'] = pre_val
        total_pre += pre_val
        row_pre += 1
    
    ws2[f'A{row_pre}'] = "TOTAL"
    ws2[f'A{row_pre}'].font = Font(bold=True)
    ws2[f'B{row_pre}'] = total_pre
    ws2[f'B{row_pre}'].font = Font(bold=True)

    chart2 = BarChart()
    chart2.title = "Salones - Precios (Mensual)"
    chart2.y_axis.title = None
    chart2.x_axis.title = None

    try:
        max_pre = max([float(s["precio"] or 0.0) for s in top_precios]) if top_precios else 0
        step_pre = _nice_step(max_pre, prefer_two=False)
        y_max_pre = int(((int(max_pre) + step_pre - 1) // step_pre) * step_pre) if step_pre else int(max_pre)
        if y_max_pre == 0:
            y_max_pre = step_pre or 1000
        chart2.y_axis.scaling.minVal = 0
        chart2.y_axis.scaling.maxVal = y_max_pre
        chart2.y_axis.majorUnit = step_pre
        chart2.y_axis.tickLblPos = "low"
        chart2.y_axis.delete = False
        chart2.y_axis.majorTickMark = "out"
        chart2.y_axis.majorGridlines.graphicalProperties.ln.solidFill = "D3D3D3"
        chart2.y_axis.majorGridlines.graphicalProperties.ln.w = 12700
    except Exception:
        pass

    try:
        chart2.x_axis.tickLblPos = "low"
        chart2.x_axis.majorTickMark = "out"
        chart2.x_axis.delete = False
        chart2.legend = None
    except Exception:
        pass

    data_ref = Reference(ws2, min_col=2, min_row=5, max_row=row_pre - 1)
    cats_ref = Reference(ws2, min_col=1, min_row=5, max_row=row_pre - 1)
    chart2.add_data(data_ref, titles_from_data=False)
    chart2.set_categories(cats_ref)

    try:
        chart2.gapWidth = 50
    except Exception:
        pass
    chart2.height = 10
    chart2.width = 18
    ws2.add_chart(chart2, "D4")

    # HOJA 3: Ingresos
    ws3 = wb.create_sheet("Ingresos")
    if fecha_descarga:
        ws3["F1"] = str(fecha_descarga)
    ws3['A1'] = f"Salones con más Ingresos - {fecha_actual.strftime('%B %Y')}"
    ws3['A1'].font = Font(bold=True, size=12)
    
    ws3['A4'] = "Salón"
    ws3['B4'] = "Ingresos (MXN)"
    ws3['A4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws3['B4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws3['A4'].font = Font(color="FFFFFF", bold=True)
    ws3['B4'].font = Font(color="FFFFFF", bold=True)
    
    row_ing = 5
    total_ing = 0.0
    for salon in top_ingresos:
        ws3[f'A{row_ing}'] = salon["salon__nombre"]
        ing_val = float(salon["total"] or 0.0)
        ws3[f'B{row_ing}'] = ing_val
        total_ing += ing_val
        row_ing += 1
    
    ws3[f'A{row_ing}'] = "TOTAL"
    ws3[f'A{row_ing}'].font = Font(bold=True)
    ws3[f'B{row_ing}'] = total_ing
    ws3[f'B{row_ing}'].font = Font(bold=True)

    chart3 = BarChart()
    chart3.title = "Salones - Ingresos (Mensual)"
    chart3.y_axis.title = None
    chart3.x_axis.title = None

    try:
        ing_vals = [float(s["total"] or 0.0) for s in top_ingresos]
        max_ing = max(ing_vals) if ing_vals else 0
        step_ing = _nice_step(max_ing, prefer_two=False)
        y_max_ing = int(((int(max_ing) + step_ing - 1) // step_ing) * step_ing) if step_ing else int(max_ing)
        if y_max_ing == 0:
            y_max_ing = step_ing or 1000
        chart3.y_axis.scaling.minVal = 0
        chart3.y_axis.scaling.maxVal = y_max_ing
        chart3.y_axis.majorUnit = step_ing
        chart3.y_axis.tickLblPos = "low"
        chart3.y_axis.delete = False
        chart3.y_axis.majorTickMark = "out"
        chart3.y_axis.majorGridlines.graphicalProperties.ln.solidFill = "D3D3D3"
        chart3.y_axis.majorGridlines.graphicalProperties.ln.w = 12700
    except Exception:
        pass

    try:
        chart3.x_axis.tickLblPos = "low"
        chart3.x_axis.majorTickMark = "out"
        chart3.x_axis.delete = False
        chart3.legend = None
    except Exception:
        pass

    data_ref = Reference(ws3, min_col=2, min_row=5, max_row=row_ing - 1)
    cats_ref = Reference(ws3, min_col=1, min_row=5, max_row=row_ing - 1)
    chart3.add_data(data_ref, titles_from_data=False)
    chart3.set_categories(cats_ref)

    try:
        chart3.gapWidth = 50
    except Exception:
        pass
    chart3.height = 10
    chart3.width = 18
    ws3.add_chart(chart3, "D4")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    from django.http import HttpResponse
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="salones_mensual.xlsx"'
    return response