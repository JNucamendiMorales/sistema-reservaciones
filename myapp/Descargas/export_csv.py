import io
import csv
import zipfile
# FIX: importar datetime además de timedelta (evita NameError)
from datetime import timedelta, datetime
from django.utils import timezone
from django.http import HttpResponse
from django.db.models import Count, Sum
from myapp.models import Salon, Reservacion
from django.contrib.auth.models import User
import pytz


# ===============================================================
# Función genérica (compatibilidad)
# ===============================================================
def exportar_csv(charts_data, images=None, fecha_descarga=None, titulo_reporte=""):
    """
    Función genérica de CSV (no se usa, mantiene compatibilidad)
    """
    return HttpResponse("Función CSV genérica no implementada", status=501)


# ===============================================================
# CSV → Excel con 3 hojas (RESERVACIONES SEMANAL)
# ===============================================================
def exportar_csv_reservaciones_semanal(fecha_descarga=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from datetime import timedelta
    from django.utils import timezone
    from django.db.models import Count, Sum

    tz = pytz.timezone('America/Mexico_City')
    hoy_local = timezone.now().astimezone(tz).date()
    inicio_semana = hoy_local - timedelta(days=hoy_local.weekday())
    dias_nombres = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]

    wb = Workbook()
    wb.remove(wb.active)

    # HOJA 1: Reservaciones por día
    ws1 = wb.create_sheet("Reservaciones")
    ws1['A1'] = "RESERVACIONES SEMANALES"
    ws1['A1'].font = Font(bold=True, size=12)
    
    if fecha_descarga:
        ws1['A2'] = f"Fecha: {fecha_descarga}"
    
    ws1['A4'] = "Día"
    ws1['B4'] = "Reservaciones"
    ws1['A4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws1['B4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws1['A4'].font = Font(color="FFFFFF", bold=True)
    ws1['B4'].font = Font(color="FFFFFF", bold=True)

    conteos = []
    for i in range(7):
        fecha = inicio_semana + timedelta(days=i)
        count = Reservacion.objects.filter(fecha_reserva=fecha).count()
        conteos.append(count)

    row = 5
    total_reservaciones = 0
    for nombre, valor in zip(dias_nombres, conteos):
        ws1[f'A{row}'] = nombre
        ws1[f'B{row}'] = valor
        total_reservaciones += valor
        row += 1

    ws1[f'A{row}'] = "TOTAL"
    ws1[f'A{row}'].font = Font(bold=True)
    ws1[f'B{row}'] = total_reservaciones
    ws1[f'B{row}'].font = Font(bold=True)

    # HOJA 2: Salones más reservados
    ws2 = wb.create_sheet("Salones")
    ws2['A1'] = "SALONES MÁS RESERVADOS"
    ws2['A1'].font = Font(bold=True, size=12)
    
    ws2['A4'] = "Salón"
    ws2['B4'] = "Reservaciones"
    ws2['A4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws2['B4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws2['A4'].font = Font(color="FFFFFF", bold=True)
    ws2['B4'].font = Font(color="FFFFFF", bold=True)

    por_salon_qs = Reservacion.objects.filter(
        fecha_reserva__range=[inicio_semana, inicio_semana + timedelta(days=6)]
    ).values("salon__nombre").annotate(total=Count("id")).order_by("-total")[:10]

    row = 5
    total_salones = 0
    for salon in por_salon_qs:
        ws2[f'A{row}'] = salon["salon__nombre"]
        ws2[f'B{row}'] = salon["total"]
        total_salones += salon["total"]
        row += 1

    ws2[f'A{row}'] = "TOTAL"
    ws2[f'A{row}'].font = Font(bold=True)
    ws2[f'B{row}'] = total_salones
    ws2[f'B{row}'].font = Font(bold=True)

    # HOJA 3: Ingresos por día
    ws3 = wb.create_sheet("Ingresos")
    ws3['A1'] = "INGRESOS SEMANALES"
    ws3['A1'].font = Font(bold=True, size=12)
    
    ws3['A4'] = "Día"
    ws3['B4'] = "Ingresos"
    ws3['A4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws3['B4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws3['A4'].font = Font(color="FFFFFF", bold=True)
    ws3['B4'].font = Font(color="FFFFFF", bold=True)

    ingresos = []
    for i in range(7):
        fecha = inicio_semana + timedelta(days=i)
        ingreso = Reservacion.objects.filter(
            fecha_reserva=fecha
        ).aggregate(Sum("precio_total"))["precio_total__sum"] or 0.0
        ingresos.append(ingreso)

    row = 5
    total_ingresos = 0.0
    for nombre, valor in zip(dias_nombres, ingresos):
        ws3[f'A{row}'] = nombre
        ws3[f'B{row}'] = float(valor)
        total_ingresos += float(valor)
        row += 1

    ws3[f'A{row}'] = "TOTAL"
    ws3[f'A{row}'].font = Font(bold=True)
    ws3[f'B{row}'] = total_ingresos
    ws3[f'B{row}'].font = Font(bold=True)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="reservaciones_semanal.xlsx"'
    wb.save(response)
    return response


# ===============================================================
# CSV → Excel con 3 hojas (RESERVACIONES MENSUAL)
# ===============================================================
def exportar_csv_reservaciones_mensual(fecha_descarga=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from django.db.models import Count, Sum

    tz = pytz.timezone('America/Mexico_City')
    fecha_actual = timezone.now().astimezone(tz)
    mes_actual = fecha_actual.month
    año_actual = fecha_actual.year

    wb = Workbook()
    wb.remove(wb.active)

    rangos = [(1, 5, "1-5"), (6, 10, "6-10"), (11, 15, "11-15"), (16, 20, "16-20"), (21, 25, "21-25"), (26, 31, "26-31")]

    # HOJA 1: Reservaciones por período
    ws1 = wb.create_sheet("Reservaciones")
    ws1['A1'] = "RESERVACIONES MENSUALES"
    ws1['A1'].font = Font(bold=True, size=12)
    if fecha_descarga:
        ws1['A2'] = f"Fecha: {fecha_descarga}"
    ws1['A4'] = "Período"
    ws1['B4'] = "Reservaciones"
    ws1['A4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws1['B4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws1['A4'].font = Font(color="FFFFFF", bold=True)
    ws1['B4'].font = Font(color="FFFFFF", bold=True)

    row = 5
    total_res = 0
    for start, end, label in rangos:
        count = Reservacion.objects.filter(
            fecha_reserva__year=año_actual,
            fecha_reserva__month=mes_actual,
            fecha_reserva__day__range=[start, end]
        ).count()
        ws1[f'A{row}'] = label
        ws1[f'B{row}'] = count
        total_res += count
        row += 1
    ws1[f'A{row}'] = "TOTAL"
    ws1[f'A{row}'].font = Font(bold=True)
    ws1[f'B{row}'] = total_res
    ws1[f'B{row}'].font = Font(bold=True)

    # HOJA 2: Salones más reservados
    ws2 = wb.create_sheet("Salones")
    ws2['A1'] = "SALONES MÁS RESERVADOS"
    ws2['A1'].font = Font(bold=True, size=12)
    ws2['A4'] = "Salón"
    ws2['B4'] = "Reservaciones"
    ws2['A4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws2['B4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws2['A4'].font = Font(color="FFFFFF", bold=True)
    ws2['B4'].font = Font(color="FFFFFF", bold=True)

    por_salon = Reservacion.objects.filter(
        fecha_reserva__year=año_actual,
        fecha_reserva__month=mes_actual
    ).values("salon__nombre").annotate(total=Count("id")).order_by("-total")[:50]

    row = 5
    total_sal = 0
    for s in por_salon:
        ws2[f'A{row}'] = s["salon__nombre"]
        ws2[f'B{row}'] = s["total"]
        total_sal += s["total"]
        row += 1
    ws2[f'A{row}'] = "TOTAL"
    ws2[f'A{row}'].font = Font(bold=True)
    ws2[f'B{row}'] = total_sal
    ws2[f'B{row}'].font = Font(bold=True)

    # HOJA 3: Ingresos por período
    ws3 = wb.create_sheet("Ingresos")
    ws3['A1'] = "INGRESOS MENSUALES"
    ws3['A1'].font = Font(bold=True, size=12)
    ws3['A4'] = "Período"
    ws3['B4'] = "Ingresos"
    ws3['A4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws3['B4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws3['A4'].font = Font(color="FFFFFF", bold=True)
    ws3['B4'].font = Font(color="FFFFFF", bold=True)

    row = 5
    total_ing = 0.0
    for start, end, label in rangos:
        ingreso = Reservacion.objects.filter(
            fecha_reserva__year=año_actual,
            fecha_reserva__month=mes_actual,
            fecha_reserva__day__range=[start, end]
        ).aggregate(Sum("precio_total"))["precio_total__sum"] or 0.0
        ws3[f'A{row}'] = label
        ws3[f'B{row}'] = float(ingreso)
        total_ing += float(ingreso)
        row += 1
    ws3[f'A{row}'] = "TOTAL"
    ws3[f'A{row}'].font = Font(bold=True)
    ws3[f'B{row}'] = total_ing
    ws3[f'B{row}'].font = Font(bold=True)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="reservaciones_mensual.xlsx"'
    wb.save(response)
    return response


# ===============================================================
# CSV → Excel con 3 hojas (USUARIOS SEMANAL)
# ===============================================================
def exportar_csv_usuarios_semanal(fecha_descarga=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from datetime import timedelta
    from django.utils import timezone
    import pytz

    tz = pytz.timezone('America/Mexico_City')
    hoy = timezone.now().astimezone(tz).date()
    inicio_semana = hoy - timedelta(days=hoy.weekday())
    dias_nombres = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]

    wb = Workbook()
    wb.remove(wb.active)

    # HOJA 1: Usuarios nuevos - CON ZONA HORARIA
    ws1 = wb.create_sheet("Nuevos Usuarios")
    ws1['A1'] = "Nuevos usuarios registrados esta semana"
    ws1['A1'].font = Font(bold=True, size=12)
    if fecha_descarga:
        ws1['A2'] = f"Fecha: {fecha_descarga}"
    ws1['A4'] = "Día"
    ws1['B4'] = "Usuarios Nuevos"
    ws1['A4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws1['B4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws1['A4'].font = Font(color="FFFFFF", bold=True)
    ws1['B4'].font = Font(color="FFFFFF", bold=True)

    row = 5
    total_nuevos = 0
    for i in range(7):
        fecha = inicio_semana + timedelta(days=i)
        fecha_inicio = timezone.make_aware(timezone.datetime.combine(fecha, timezone.datetime.min.time()), tz)
        fecha_fin = timezone.make_aware(timezone.datetime.combine(fecha, timezone.datetime.max.time()), tz)
        count = User.objects.filter(date_joined__gte=fecha_inicio, date_joined__lte=fecha_fin).count()
        ws1[f'A{row}'] = dias_nombres[i]
        ws1[f'B{row}'] = count
        total_nuevos += count
        row += 1
    ws1[f'A{row}'] = "TOTAL"
    ws1[f'A{row}'].font = Font(bold=True)
    ws1[f'B{row}'] = total_nuevos
    ws1[f'B{row}'].font = Font(bold=True)

    # HOJA 2: Usuarios activos
    ws2 = wb.create_sheet("Usuarios Activos")
    ws2['A1'] = "Usuarios con mas Reservaciones esta semana"
    ws2['A1'].font = Font(bold=True, size=12)
    ws2['A4'] = "Usuario"
    ws2['B4'] = "Reservaciones"
    ws2['A4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws2['B4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws2['A4'].font = Font(color="FFFFFF", bold=True)
    ws2['B4'].font = Font(color="FFFFFF", bold=True)

    usuarios_activos = Reservacion.objects.filter(
        fecha_reserva__range=[inicio_semana, inicio_semana + timedelta(days=6)]
    ).values("usuario__username").annotate(total=Count("id")).order_by("-total")[:50]

    row = 5
    total_activos = 0
    for u in usuarios_activos:
        ws2[f'A{row}'] = u["usuario__username"]
        ws2[f'B{row}'] = u["total"]
        total_activos += u["total"]
        row += 1
    ws2[f'A{row}'] = "TOTAL"
    ws2[f'A{row}'].font = Font(bold=True)
    ws2[f'B{row}'] = total_activos
    ws2[f'B{row}'].font = Font(bold=True)

    # HOJA 3: Ingresos por usuario
    ws3 = wb.create_sheet("Ingresos")
    ws3['A1'] = "Usuarios con mayores ingresos esta semana"
    ws3['A1'].font = Font(bold=True, size=12)
    ws3['A4'] = "Usuario"
    ws3['B4'] = "Ingresos"
    ws3['A4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws3['B4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws3['A4'].font = Font(color="FFFFFF", bold=True)
    ws3['B4'].font = Font(color="FFFFFF", bold=True)

    usuarios_ingresos = Reservacion.objects.filter(
        fecha_reserva__range=[inicio_semana, inicio_semana + timedelta(days=6)]
    ).values("usuario__username").annotate(total=Sum("precio_total")).order_by("-total")[:50]

    row = 5
    total_ing = 0.0
    for u in usuarios_ingresos:
        ws3[f'A{row}'] = u["usuario__username"]
        ws3[f'B{row}'] = float(u["total"] or 0.0)
        total_ing += float(u["total"] or 0.0)
        row += 1
    ws3[f'A{row}'] = "TOTAL"
    ws3[f'A{row}'].font = Font(bold=True)
    ws3[f'B{row}'] = total_ing
    ws3[f'B{row}'].font = Font(bold=True)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="usuarios_semanal.xlsx"'
    wb.save(response)
    return response


# ===============================================================
# CSV → Excel con 3 hojas (USUARIOS MENSUAL)
# ===============================================================
def exportar_csv_usuarios_mensual(fecha_descarga=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from django.utils import timezone
    import pytz

    tz = pytz.timezone('America/Mexico_City')
    fecha_actual = timezone.now().astimezone(tz)
    mes_actual = fecha_actual.month
    año_actual = fecha_actual.year

    wb = Workbook()
    wb.remove(wb.active)

    rangos = [(1, 5, "1-5"), (6, 10, "6-10"), (11, 15, "11-15"), (16, 20, "16-20"), (21, 25, "21-25"), (26, 31, "26-31")]

    # HOJA 1: Nuevos usuarios - CON ZONA HORARIA
    ws1 = wb.create_sheet("Nuevos Usuarios")
    ws1['A1'] = "Nuevos usuarios registrados este mes"
    ws1['A1'].font = Font(bold=True, size=12)
    if fecha_descarga:
        ws1['A2'] = f"Fecha: {fecha_descarga}"
    ws1['A4'] = "Período"
    ws1['B4'] = "Usuarios Nuevos"
    ws1['A4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws1['B4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws1['A4'].font = Font(color="FFFFFF", bold=True)
    ws1['B4'].font = Font(color="FFFFFF", bold=True)

    row = 5
    total_nuevos = 0
    for start, end, label in rangos:
        fecha_inicio = timezone.make_aware(timezone.datetime(año_actual, mes_actual, start, 0, 0, 0), tz)
        fecha_fin = timezone.make_aware(timezone.datetime(año_actual, mes_actual, min(end, 31), 23, 59, 59), tz)
        count = User.objects.filter(date_joined__gte=fecha_inicio, date_joined__lte=fecha_fin).count()
        ws1[f'A{row}'] = label
        ws1[f'B{row}'] = count
        total_nuevos += count
        row += 1
    ws1[f'A{row}'] = "TOTAL"
    ws1[f'A{row}'].font = Font(bold=True)
    ws1[f'B{row}'] = total_nuevos
    ws1[f'B{row}'].font = Font(bold=True)

    # HOJA 2: Usuarios activos
    ws2 = wb.create_sheet("Usuarios Activos")
    ws2['A1'] = "Usuarios con mas Reservaciones este mes"
    ws2['A1'].font = Font(bold=True, size=12)
    ws2['A4'] = "Usuario"
    ws2['B4'] = "Reservaciones"
    ws2['A4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws2['B4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws2['A4'].font = Font(color="FFFFFF", bold=True)
    ws2['B4'].font = Font(color="FFFFFF", bold=True)

    usuarios_activos = Reservacion.objects.filter(
        fecha_reserva__year=año_actual,
        fecha_reserva__month=mes_actual
    ).values("usuario__username").annotate(total=Count("id")).order_by("-total")[:50]

    row = 5
    total_activos = 0
    for u in usuarios_activos:
        ws2[f'A{row}'] = u["usuario__username"]
        ws2[f'B{row}'] = u["total"]
        total_activos += u["total"]
        row += 1
    ws2[f'A{row}'] = "TOTAL"
    ws2[f'A{row}'].font = Font(bold=True)
    ws2[f'B{row}'] = total_activos
    ws2[f'B{row}'].font = Font(bold=True)

    # HOJA 3: Ingresos por usuario
    ws3 = wb.create_sheet("Ingresos")
    ws3['A1'] = "Usuarios con mayores ingresos este mes"
    ws3['A1'].font = Font(bold=True, size=12)
    ws3['A4'] = "Usuario"
    ws3['B4'] = "Ingresos"
    ws3['A4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws3['B4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws3['A4'].font = Font(color="FFFFFF", bold=True)
    ws3['B4'].font = Font(color="FFFFFF", bold=True)

    usuarios_ingresos = Reservacion.objects.filter(
        fecha_reserva__year=año_actual,
        fecha_reserva__month=mes_actual
    ).values("usuario__username").annotate(total=Sum("precio_total")).order_by("-total")[:50]

    row = 5
    total_ing = 0.0
    for u in usuarios_ingresos:
        ws3[f'A{row}'] = u["usuario__username"]
        ws3[f'B{row}'] = float(u["total"] or 0.0)
        total_ing += float(u["total"] or 0.0)
        row += 1
    ws3[f'A{row}'] = "TOTAL"
    ws3[f'A{row}'].font = Font(bold=True)
    ws3[f'B{row}'] = total_ing
    ws3[f'B{row}'].font = Font(bold=True)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="usuarios_mensual.xlsx"'
    wb.save(response)
    return response


# ===============================================================
# CSV → Excel con 3 hojas (SALONES SEMANAL)
# ===============================================================
def exportar_csv_salones_semanal(fecha_descarga=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from django.utils import timezone
    import pytz

    tz = pytz.timezone('America/Mexico_City')
    hoy = timezone.now().astimezone(tz).date()
    inicio_semana = hoy - timedelta(days=hoy.weekday())
    dias_nombres = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]

    wb = Workbook()
    wb.remove(wb.active)

    # HOJA 1: Calificación
    ws1 = wb.create_sheet("Calificación")
    ws1['A1'] = "Salones con mejores reseñas (Semanal)"
    ws1['A1'].font = Font(bold=True, size=12)
    if fecha_descarga:
        ws1['A2'] = f"Fecha: {fecha_descarga}"

    ws1['A4'] = "Salón"
    ws1['B4'] = "Calificación"
    ws1['A4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws1['B4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws1['A4'].font = Font(color="FFFFFF", bold=True)
    ws1['B4'].font = Font(color="FFFFFF", bold=True)

    top_cal = list(Salon.objects.values("nombre", "calificacion").order_by("-calificacion")[:5])
    row = 5
    total_cal = 0.0
    for s in top_cal:
        ws1[f'A{row}'] = s["nombre"]
        ws1[f'B{row}'] = float(s.get("calificacion") or 0.0)
        total_cal += float(s.get("calificacion") or 0.0)
        row += 1

    # REMOVED: TOTAL row for calificacion (do not sum ratings)
    # previously added TOTAL row here

    # HOJA 2: Precios
    ws2 = wb.create_sheet("Precios")
    ws2['A1'] = "Salones más caros (Semanal)"
    ws2['A1'].font = Font(bold=True, size=12)
    if fecha_descarga:
        ws2['A2'] = f"Fecha: {fecha_descarga}"
    ws2['A4'] = "Salón"
    ws2['B4'] = "Precio (MXN)"
    ws2['A4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws2['B4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws2['A4'].font = Font(color="FFFFFF", bold=True)
    ws2['B4'].font = Font(color="FFFFFF", bold=True)

    top_price = list(Salon.objects.values("nombre", "precio").order_by("-precio")[:5])
    row = 5
    total_price = 0.0
    for s in top_price:
        ws2[f'A{row}'] = s["nombre"]
        ws2[f'B{row}'] = float(s.get("precio") or 0.0)
        total_price += float(s.get("precio") or 0.0)
        row += 1
    ws2[f'A{row}'] = "TOTAL"
    ws2[f'B{row}'] = total_price
    ws2[f'A{row}'].font = Font(bold=True)
    ws2[f'B{row}'].font = Font(bold=True)

    # HOJA 3: Ingresos por salón (semana)
    ws3 = wb.create_sheet("Ingresos")
    ws3['A1'] = "Salones con más ingresos (Semanal)"
    ws3['A1'].font = Font(bold=True, size=12)
    if fecha_descarga:
        ws3['A2'] = f"Fecha: {fecha_descarga}"
    ws3['A4'] = "Salón"
    ws3['B4'] = "Ingresos (MXN)"
    ws3['A4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws3['B4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws3['A4'].font = Font(color="FFFFFF", bold=True)
    ws3['B4'].font = Font(color="FFFFFF", bold=True)

    start_dt = timezone.make_aware(datetime.combine(inicio_semana, datetime.min.time()), tz)
    end_dt = timezone.make_aware(datetime.combine(inicio_semana + timedelta(days=6), datetime.max.time()), tz)
    por_salon_ing = Reservacion.objects.filter(fecha_reserva__range=[start_dt, end_dt]) \
        .values("salon__nombre").annotate(total=Sum("precio_total")).order_by("-total")[:10]

    row = 5
    total_ing = 0.0
    for s in por_salon_ing:
        ws3[f'A{row}'] = s["salon__nombre"]
        val = float(s.get("total") or 0.0)
        ws3[f'B{row}'] = val
        total_ing += val
        row += 1
    ws3[f'A{row}'] = "TOTAL"
    ws3[f'B{row}'] = total_ing
    ws3[f'A{row}'].font = Font(bold=True)
    ws3[f'B{row}'].font = Font(bold=True)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="salones_semanal.xlsx"'
    wb.save(response)
    return response


# ===============================================================
# CSV → Excel con 3 hojas (SALONES MENSUAL)
# ===============================================================
def exportar_csv_salones_mensual(fecha_descarga=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from django.utils import timezone
    import pytz
    import calendar

    tz = pytz.timezone('America/Mexico_City')
    fecha_actual = timezone.now().astimezone(tz)
    mes_actual = fecha_actual.month
    año_actual = fecha_actual.year

    wb = Workbook()
    wb.remove(wb.active)

    rangos = [(1,5,"1-5"), (6,10,"6-10"), (11,15,"11-15"), (16,20,"16-20"), (21,25,"21-25"), (26,31,"26-31")]

    # HOJA 1: Calificación (top 5)
    ws1 = wb.create_sheet("Calificación")
    ws1['A1'] = f"Salones con mejores reseñas - {fecha_actual.strftime('%B %Y')}"
    ws1['A1'].font = Font(bold=True, size=12)
    if fecha_descarga:
        ws1['A2'] = f"Fecha: {fecha_descarga}"
    ws1['A4'] = "Salón"
    ws1['B4'] = "Calificación"
    ws1['A4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws1['B4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws1['A4'].font = Font(color="FFFFFF", bold=True)
    ws1['B4'].font = Font(color="FFFFFF", bold=True)

    top_cal = list(Salon.objects.values("nombre","calificacion").order_by("-calificacion")[:5])
    row = 5
    total_cal = 0.0
    for s in top_cal:
        ws1[f'A{row}'] = s["nombre"]
        ws1[f'B{row}'] = float(s.get("calificacion") or 0.0)
        total_cal += float(s.get("calificacion") or 0.0)
        row += 1
    ws1[f'A{row}'] = "TOTAL"
    ws1[f'B{row}'] = total_cal
    ws1[f'A{row}'].font = Font(bold=True)
    ws1[f'B{row}'].font = Font(bold=True)

    # HOJA 2: Precios
    ws2 = wb.create_sheet("Precios")
    ws2['A1'] = f"Salones más caros - {fecha_actual.strftime('%B %Y')}"
    ws2['A1'].font = Font(bold=True, size=12)
    ws2['A4'] = "Salón"
    ws2['B4'] = "Precio (MXN)"
    ws2['A4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws2['B4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws2['A4'].font = Font(color="FFFFFF", bold=True)
    ws2['B4'].font = Font(color="FFFFFF", bold=True)

    top_price = list(Salon.objects.values("nombre","precio").order_by("-precio")[:5])
    row = 5
    total_price = 0.0
    for s in top_price:
        ws2[f'A{row}'] = s["nombre"]
        ws2[f'B{row}'] = float(s.get("precio") or 0.0)
        total_price += float(s.get("precio") or 0.0)
        row += 1
    ws2[f'A{row}'] = "TOTAL"
    ws2[f'B{row}'] = total_price
    ws2[f'A{row}'].font = Font(bold=True)
    ws2[f'B{row}'].font = Font(bold=True)

    # HOJA 3: Ingresos mensual por rangos
    ws3 = wb.create_sheet("Ingresos")
    ws3['A1'] = f"Salones con más ingresos - {fecha_actual.strftime('%B %Y')}"
    ws3['A1'].font = Font(bold=True, size=12)
    ws3['A4'] = "Salón"
    ws3['B4'] = "Ingresos (MXN)"
    ws3['A4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws3['B4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws3['A4'].font = Font(color="FFFFFF", bold=True)
    ws3['B4'].font = Font(color="FFFFFF", bold=True)

    # Acumular ingresos por salón para mes
    sal_ing = {}
    last_day = calendar.monthrange(año_actual, mes_actual)[1]
    for start, end, label in rangos:
        end_day = min(end, last_day)
        qs = Reservacion.objects.filter(
            fecha_reserva__year=año_actual,
            fecha_reserva__month=mes_actual,
            fecha_reserva__day__gte=start,
            fecha_reserva__day__lte=end_day
        ).values("salon__nombre").annotate(total=Sum("precio_total"))
        for it in qs:
            name = it["salon__nombre"]
            sal_ing[name] = sal_ing.get(name, 0.0) + float(it["total"] or 0.0)

    # ordenar por total
    sorted_sal = sorted(sal_ing.items(), key=lambda x: x[1], reverse=True)[:10]
    row = 5
    total_ing = 0.0
    for name, val in sorted_sal:
        ws3[f'A{row}'] = name
        ws3[f'B{row}'] = float(val)
        total_ing += float(val)
        row += 1
    ws3[f'A{row}'] = "TOTAL"
    ws3[f'B{row}'] = total_ing
    ws3[f'A{row}'].font = Font(bold=True)
    ws3[f'B{row}'].font = Font(bold=True)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="salones_mensual.xlsx"'
    wb.save(response)
    return response


# ===============================================================
# VISTAS / HELPERS
# ===============================================================
def generar_datos_para(charts):
    """Genera datos base para gráficas"""
    base_data = [
        {"Etiqueta": "Lunes", "Valor": 12},
        {"Etiqueta": "Martes", "Valor": 18},
        {"Etiqueta": "Miércoles", "Valor": 9},
        {"Etiqueta": "Jueves", "Valor": 14},
        {"Etiqueta": "Viernes", "Valor": 11},
        {"Etiqueta": "Sábado", "Valor": 16},
        {"Etiqueta": "Domingo", "Valor": 8},
    ]

    charts_data = []
    for name in charts:
        charts_data.append({
            "nombre": name.replace("_", " ").capitalize(),
            "datos": base_data
        })
    return charts_data


def descargar_csv(request):
    """Vista para descargar CSV desde request"""
    charts = request.GET.get("charts", "").split(",")
    return exportar_csv(generar_datos_para(charts))






