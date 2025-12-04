import io
import base64
import csv
from datetime import timedelta

# asegurarse de tener ImageReader importado (ya estaba)
from reportlab.lib.utils import ImageReader

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Font

from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from django.http import HttpResponse
from myapp.models import Salon, Reservacion


# ===============================================================
# CSV
# ===============================================================
def exportar_csv(charts_data, fecha_descarga=None):
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Agregar fecha en la primera fila
    if fecha_descarga:
        writer.writerow(["Fecha de descarga:", fecha_descarga])
        writer.writerow([])

    for chart in charts_data:
        writer.writerow([chart.get("nombre", "")])
        writer.writerow(["Etiqueta", "Valor"])
        for fila in chart.get("datos", []):
            writer.writerow([fila.get("Etiqueta"), fila.get("Valor")])
        writer.writerow([])

    response = HttpResponse(output.getvalue(), content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="graficas.csv"'
    return response


# ===============================================================
# XLSX
# ===============================================================
def exportar_xlsx_nativo(charts_data, fecha_descarga=None):
    wb = Workbook()
    ws_default = wb.active
    wb.remove(ws_default)

    for chart in charts_data:

        ws = wb.create_sheet(title=chart["nombre"][:31])
        ws.append(["Etiqueta", "Valor"])

        for fila in chart["datos"]:
            ws.append([fila.get("Etiqueta", ""), fila.get("Valor", "")])

        # Selección automática del gráfico
        name = chart["nombre"].lower()
        if "line" in name or "línea" in name:
            c = LineChart()
        elif "pie" in name or "pastel" in name:
            c = PieChart()
        else:
            c = BarChart()

        c.title = chart["nombre"]

        data = Reference(ws, min_col=2, min_row=2, max_row=len(chart["datos"]) + 1)
        categories = Reference(ws, min_col=1, min_row=2, max_row=len(chart["datos"]) + 1)

        c.add_data(data, titles_from_data=False)
        c.set_categories(categories)

        ws.add_chart(c, "D2")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="graficas.xlsx"'
    return response



# ===============================================================
# XLSX — Reservaciones por periodo (SEMANAL) con ejes correctos
# ===============================================================
def exportar_xlsx_reservaciones_semanal(fecha_descarga=None):
    from django.utils import timezone
    from django.db.models import Count
    import io
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Font

    hoy = timezone.now().date()
    inicio_semana = hoy - timedelta(days=hoy.weekday())
    dias_nombres = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]

    # Conteo por día (consultas simples por fecha - fecha_reserva es DateField)
    conteos = []
    for i in range(7):
        d = inicio_semana + timedelta(days=i)
        total = Reservacion.objects.filter(fecha_reserva=d).count()
        conteos.append(total)

    wb = Workbook()
    ws = wb.active
    ws.title = "Reservaciones por periodo"

    # Fecha en esquina superior derecha
    if fecha_descarga:
        ws["H1"] = "Fecha:"
        ws["I1"] = fecha_descarga
        ws["H1"].font = Font(bold=True)

    # Cabecera y filas
    ws.append(["Reservaciones por periodo (Semanal)"])
    ws.append([])
    ws.append(["Día", "Reservaciones"])
    for nombre, valor in zip(dias_nombres, conteos):
        ws.append([nombre, valor])

    # Crear gráfico de barras
    chart = BarChart()
    chart.title = "Reservaciones - Semana"
    chart.y_axis.title = "Reservaciones"
    chart.x_axis.title = "Día"

    data_ref = Reference(ws, min_col=2, min_row=4, max_row=4+len(conteos)-1)
    cats_ref = Reference(ws, min_col=1, min_row=4, max_row=4+len(conteos)-1)
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cats_ref)
    chart.height = 8
    chart.width = 16

    ws.add_chart(chart, "D4")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="reservaciones_semanal.xlsx"'
    return response
    

#Exportar mensual    
    
def exportar_xlsx_reservaciones_mensual(fecha_descarga=None):
    from django.db.models import Count
    from django.db.models.functions import ExtractDay
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Font
    from django.http import HttpResponse
    from myapp.models import Reservacion
    from django.utils.timezone import now

    fecha_actual = now()
    mes_actual = fecha_actual.month
    año_actual = fecha_actual.year

    # Rangos de días
    dias_rango = [
        (1, 5, "1-5"),
        (6, 10, "6-10"),
        (11, 15, "11-15"),
        (16, 20, "16-20"),
        (21, 25, "21-25"),
        (26, 31, "26-31"),
    ]

    # Query: Reservaciones del mes actual
    data = (
        Reservacion.objects
        .filter(fecha_reserva__year=año_actual, fecha_reserva__month=mes_actual)
        .annotate(dia=ExtractDay("fecha_reserva"))
        .values("dia")
        .annotate(total=Count("id"))
    )

    # Convertir a dict {dia: total}
    conteo_por_dia = {d["dia"]: d["total"] for d in data}

    # Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Reservaciones Mensuales"
    
    # Agregar fecha en esquina superior derecha
    if fecha_descarga:
        ws["H1"] = "Fecha:"
        ws["I1"] = fecha_descarga
        ws["H1"].font = Font(bold=True)

    ws.append([f"Reservaciones por periodo - {fecha_actual.strftime('%B %Y')}"])
    ws.append([])
    ws.append(["Días", "Reservaciones"])

    # Sumar reservas por rango
    for start, end, label in dias_rango:
        total = sum(conteo_por_dia.get(d, 0) for d in range(start, end + 1))
        ws.append([label, total])

    # Chart
    chart = BarChart()
    chart.type = "col"
    chart.style = 2
    chart.title = "Reservaciones por periodo (Mensual)"

    data_ref = Reference(ws, min_col=2, min_row=4, max_row=9)
    categories_ref = Reference(ws, min_col=1, min_row=4, max_row=9)

    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(categories_ref)

    chart.x_axis.tickLblPos = "low"
    chart.x_axis.majorTickMark = "out"
    chart.x_axis.crosses = "min"

    chart.y_axis.tickLblPos = "nextTo"
    chart.y_axis.majorTickMark = "out"
    chart.y_axis.crosses = "min"

    chart.gapWidth = 180

    ws.add_chart(chart, "D4")

    # Export
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="reservaciones_{fecha_actual.strftime("%m_%Y")}.xlsx"'}
    )



# ===============================================================
# CSV — Reservaciones por periodo (SEMANAL)
# ===============================================================
def exportar_csv_reservaciones_semanal(fecha_descarga=None):
    from django.utils import timezone
    from django.db.models import Count, Sum
    import io
    import csv

    hoy = timezone.now().date()
    inicio_semana = hoy - timedelta(days=hoy.weekday())
    dias_nombres = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]

    # --- TABLA 1: Conteo por día (Reservaciones) ---
    conteos = []
    for i in range(7):
        d = inicio_semana + timedelta(days=i)
        total = Reservacion.objects.filter(fecha_reserva=d).count()
        conteos.append(total)

    # --- TABLA 2: Salones más reservados ---
    por_salon_qs = Reservacion.objects.values("salon__nombre").annotate(total=Count("id")).order_by("-total")[:10]
    datos_salones = list(por_salon_qs)

    # --- TABLA 3: Ingresos por día ---
    ingresos = []
    for i in range(7):
        d = inicio_semana + timedelta(days=i)
        total_ingreso = (Reservacion.objects
                         .filter(fecha_reserva=d)
                         .aggregate(total=Sum("precio_total"))["total"] or 0.0)
        ingresos.append(float(total_ingreso))

    wb = Workbook()
    wb.remove(wb.active)

    # ===============================================================
    # HOJA 1: Reservaciones Semanales
    # ===============================================================
    ws1 = wb.create_sheet("Reservaciones")
    
    if fecha_descarga:
        ws1["H1"] = "Fecha:"
        ws1["I1"] = fecha_descarga
        ws1["H1"].font = Font(bold=True)

    ws1.append(["Número de reservaciones semanalmente"])
    ws1.append([])
    ws1.append(["Día", "Reservaciones"])
    total_reservaciones = 0
    for nombre, valor in zip(dias_nombres, conteos):
        ws1.append([nombre, valor])
        total_reservaciones += valor
    ws1.append(["TOTAL", total_reservaciones])

    # ===============================================================
    # HOJA 2: Salones Semanales
    # ===============================================================
    ws2 = wb.create_sheet("Salones")
    
    if fecha_descarga:
        ws2["H1"] = "Fecha:"
        ws2["I1"] = fecha_descarga
        ws2["H1"].font = Font(bold=True)

    ws2.append(["Salones más reservados semanalmente"])
    ws2.append([])
    ws2.append(["Salón", "Reservaciones"])
    total_salones = 0
    for salon in datos_salones:
        ws2.append([salon["salon__nombre"], salon["total"]])
        total_salones += salon["total"]
    ws2.append(["TOTAL", total_salones])

    # ===============================================================
    # HOJA 3: Ingresos Semanales
    # ===============================================================
    ws3 = wb.create_sheet("Ingresos")
    
    if fecha_descarga:
        ws3["H1"] = "Fecha:"
        ws3["I1"] = fecha_descarga
        ws3["H1"].font = Font(bold=True)

    ws3.append(["Ingresos semanalmente"])
    ws3.append([])
    ws3.append(["Día", "Ingresos"])
    total_ingresos = 0.0
    for nombre, valor in zip(dias_nombres, ingresos):
        ws3.append([nombre, valor])
        total_ingresos += valor
    ws3.append(["TOTAL", total_ingresos])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="reservaciones_semanal.xlsx"'
    return response


# ===============================================================
# CSV — Reservaciones por periodo (MENSUAL) con Sheets separados
# ===============================================================
def exportar_csv_reservaciones_mensual(fecha_descarga=None):
    from django.utils import timezone
    from django.db.models import Count, Sum
    import io

    hoy = timezone.now().date()
    mes_actual = hoy.month
    año_actual = hoy.year

    # Rangos de días del mes
    dias_rango = [
        (1, 5, "1-5"),
        (6, 10, "6-10"),
        (11, 15, "11-15"),
        (16, 20, "16-20"),
        (21, 25, "21-25"),
        (26, 31, "26-31"),
    ]

    # --- TABLA 1: Conteo por rango de días (Reservaciones) ---
    conteos = []
    for start, end, label in dias_rango:
        total = Reservacion.objects.filter(
            fecha_reserva__year=año_actual,
            fecha_reserva__month=mes_actual,
            fecha_reserva__day__gte=start,
            fecha_reserva__day__lte=end
        ).count()
        conteos.append(total)

    # --- TABLA 2: Salones más reservados (todo el mes) ---
    por_salon_qs = Reservacion.objects.filter(
        fecha_reserva__year=año_actual,
        fecha_reserva__month=mes_actual
    ).values("salon__nombre").annotate(total=Count("id")).order_by("-total")[:10]
    datos_salones = list(por_salon_qs)

    # --- TABLA 3: Ingresos por rango de días ---
    ingresos = []
    for start, end, label in dias_rango:
        total_ingreso = (Reservacion.objects
                         .filter(
                             fecha_reserva__year=año_actual,
                             fecha_reserva__month=mes_actual,
                             fecha_reserva__day__gte=start,
                             fecha_reserva__day__lte=end
                         )
                         .aggregate(total=Sum("precio_total"))["total"] or 0.0)
        ingresos.append(float(total_ingreso))

    wb = Workbook()
    wb.remove(wb.active)

    # ===============================================================
    # HOJA 1: Reservaciones Mensuales
    # ===============================================================
    ws1 = wb.create_sheet("Reservaciones")
    
    if fecha_descarga:
        ws1["H1"] = "Fecha:"
        ws1["I1"] = fecha_descarga
        ws1["H1"].font = Font(bold=True)

    ws1.append(["Número de reservaciones mensualmente"])
    ws1.append([])
    ws1.append(["Días", "Reservaciones"])
    total_reservaciones = 0
    for label, valor in zip([r[2] for r in dias_rango], conteos):
        ws1.append([label, valor])
        total_reservaciones += valor
    ws1.append(["TOTAL", total_reservaciones])

    # ===============================================================
    # HOJA 2: Salones Mensuales
    # ===============================================================
    ws2 = wb.create_sheet("Salones")
    
    if fecha_descarga:
        ws2["H1"] = "Fecha:"
        ws2["I1"] = fecha_descarga
        ws2["H1"].font = Font(bold=True)

    ws2.append(["Salones más reservados mensualmente"])
    ws2.append([])
    ws2.append(["Salón", "Reservaciones"])
    total_salones = 0
    for salon in datos_salones:
        ws2.append([salon["salon__nombre"], salon["total"]])
        total_salones += salon["total"]
    ws2.append(["TOTAL", total_salones])

    # ===============================================================
    # HOJA 3: Ingresos Mensuales
    # ===============================================================
    ws3 = wb.create_sheet("Ingresos")
    
    if fecha_descarga:
        ws3["H1"] = "Fecha:"
        ws3["I1"] = fecha_descarga
        ws3["H1"].font = Font(bold=True)

    ws3.append(["Ingresos mensualmente"])
    ws3.append([])
    ws3.append(["Días", "Ingresos"])
    total_ingresos = 0.0
    for label, valor in zip([r[2] for r in dias_rango], ingresos):
        ws3.append([label, valor])
        total_ingresos += valor
    ws3.append(["TOTAL", total_ingresos])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="reservaciones_mensual.xlsx"'
    return response



# ===============================================================
# PDF
# ===============================================================
def exportar_pdf(charts_data, images, fecha_descarga=None, titulo_reporte="Reporte Semanal de Reservaciones"):
    """
    PDF en tema oscuro.
    - Título principal: configurable (Reporte Semanal o Mensual)
    - Cada chart se coloca en su propia página
    - Usa nombres amigables: "Reservaciones", "Salones mas reservados", "Ingresos"
    - Dibuja tablas con líneas tipo celdas
    - Omite charts que contengan 'usuario'
    - images: dict obligatorio con imágenes base64; claves normalizadas
    """
    import io
    import base64
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    from reportlab.lib.utils import ImageReader
    import unicodedata
    import re

    if not isinstance(images, dict) or not images:
        raise ValueError("El parámetro 'images' es obligatorio y debe ser un dict con imágenes base64.")

    # Filtrar charts irrelevantes (usuarios)
    charts_data = [c for c in charts_data if 'usuario' not in c.get('nombre', '').lower()]

    def normalize_key(s):
        s = s.lower()
        s = unicodedata.normalize("NFKD", s)
        s = "".join(ch for ch in s if not unicodedata.combining(ch))
        s = re.sub(r"[^\w\s]", "", s)
        return re.sub(r"\s+", "_", s).strip("_")

    # mapping display title
    def display_title_from(nombre):
        ln = nombre.lower()
        if "reservac" in ln:
            return "Reservaciones"
        if "salon" in ln or "salones" in ln or "uso de salones" in ln:
            return "Salones mas reservados"
        if "ingres" in ln or "ingreso" in ln:
            return "Ingresos"
        return nombre

    # validar imágenes para charts que quedan (usar normalized display title as key)
    for chart in charts_data:
        key = normalize_key(display_title_from(chart.get("nombre", "")))
        if key not in images or not images[key]:
            raise ValueError(f"Falta imagen para el chart '{chart.get('nombre')}' (clave esperada: '{key}').")

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    def _new_page_with_background():
        # fondo oscuro completo y encabezado
        c.setFillColor(colors.HexColor("#0f172a"))
        c.rect(0, 0, width, height, fill=1, stroke=0)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 16)
        c.drawCentredString(width / 2.0, height - 36, titulo_reporte)  # usar el título dinámico
        if fecha_descarga:
            c.setFont("Helvetica", 9)
            c.drawRightString(width - 40, height - 52, fecha_descarga)

    # dibujar cada chart en su propia página
    first = True
    left_margin = 50
    right_limit = width - 50
    col2_x = right_limit - 90
    row_height = 16
    # Ajusta estos valores a tu gusto (unidades en puntos PDF)
    # Para que la fila TOTAL quede contigua a la última fila de datos (como en Excel),
    # dejar ambos espaciamientos en 0.
    spacing_before_total = 0
    spacing_after_total = 0

    for chart in charts_data:
        if not first:
            c.showPage()
        _new_page_with_background()
        first = False

        nombre_original = chart.get("nombre", "")
        titulo = display_title_from(nombre_original)
        datos = chart.get("datos", [])

        y = height - 90

        # Subtítulo (usar titulo ya mapeado)
        c.setFont("Helvetica-Bold", 12)
        c.setFillColor(colors.whitesmoke)
        c.drawString(left_margin, y, titulo)
        y -= 28

        # Cabecera de tabla
        header_h = row_height
        c.setFillColor(colors.HexColor("#071025"))
        c.rect(left_margin - 6, y - 4, (right_limit - left_margin) + 12, header_h + 6, fill=1, stroke=0)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 10)
        
        # Headers dinámicos según el chart
        if "reservac" in titulo.lower():
            header_col1 = "Día"
            header_col2 = "Reservaciones"
        elif "salon" in titulo.lower():
            header_col1 = "Salón"
            header_col2 = "Reservaciones"
        elif "ingres" in titulo.lower():
            header_col1 = "Día"
            header_col2 = "Ingresos"
        else:
            header_col1 = "Etiqueta"
            header_col2 = "Valor"
        
        c.drawString(left_margin, y + 2, header_col1)
        c.drawString(col2_x + 6, y + 2, header_col2)

        # Cabecera bordes
        c.setStrokeColor(colors.grey)
        top_y = y + header_h + 2
        left_x = left_margin - 6
        right_x = right_limit + 6
        c.setLineWidth(0.5)
        c.line(left_x, top_y, right_x, top_y)
        c.line(left_x, y - 6, right_x, y - 6)
        c.line(col2_x, top_y, col2_x, y - 6)

        y -= (header_h + 10)

        # Filas con líneas (simulando celdas)
        c.setFont("Helvetica", 10)
        c.setFillColor(colors.lightgrey)

        total_numeric = 0.0
        rows_drawn = 0
        for fila in datos:
            etiqueta = str(fila.get("Etiqueta", ""))
            valor_raw = fila.get("Valor", 0)
            # acumular total numérico si el valor es numérico
            try:
                num_val = float(valor_raw or 0)
            except Exception:
                num_val = 0.0
            total_numeric += num_val

            # formatear para mostrar
            if "ingres" in titulo.lower():
                try:
                    valor = "${:,.2f}".format(num_val)
                except Exception:
                    valor = str(valor_raw)
            else:
                # mostrar entero para conteos
                try:
                    valor = str(int(num_val))
                except Exception:
                    valor = str(valor_raw)

            c.drawString(left_margin + 2, y, etiqueta)
            c.drawRightString(right_limit, y, valor)

            # líneas horizontales y verticales para celdas
            line_y = y - 4
            c.setStrokeColor(colors.HexColor("#203040"))
            c.setLineWidth(0.4)
            c.line(left_x, line_y, right_x, line_y)
            c.line(left_x, y + row_height, left_x, line_y)
            c.line(col2_x, y + row_height, col2_x, line_y)
            c.line(right_x, y + row_height, right_x, line_y)

            y -= row_height
            rows_drawn += 1
            if y < 260:
                break

        # Agregar fila TOTAL sólo para Reservaciones y para Ingresos
        if ("reservac" in titulo.lower()) or ("ingres" in titulo.lower()):
            # Dejar un pequeño espacio extra antes del total para evitar solapamiento
            y -= spacing_before_total

            # si no hay espacio suficiente, nueva página
            if y < 120:
                c.showPage()
                _new_page_with_background()
                y = height - 90
                # re-dibujar subtítulo y cabecera si es necesario (opcional)

            # Nota: removida la línea fina que causaba solapamiento con la última fila de datos.
            # Si se necesita una separación visible, ajustar aquí con otro estilo/espaciado.

            # Total label y valor
            c.setFont("Helvetica-Bold", 10)
            c.setFillColor(colors.whitesmoke)
            total_label = "TOTAL"
            if "ingres" in titulo.lower():
                total_display = "${:,.2f}".format(total_numeric)
            else:
                total_display = str(int(total_numeric))

            c.drawString(left_margin + 2, y, total_label)
            c.drawRightString(right_limit, y, total_display)

            # dibujar línea inferior de la fila total
            line_y = y - 4
            c.setStrokeColor(colors.HexColor("#203040"))
            c.setLineWidth(0.6)
            c.line(left_x, line_y, right_x, line_y)
            c.line(left_x, y + row_height, left_x, line_y)
            c.line(col2_x, y + row_height, col2_x, line_y)
            c.line(right_x, y + row_height, right_x, line_y)

            # bajar un poco más después del total (ajustable)
            y -= (row_height + spacing_after_total)

        # insertar imagen asociada (usar clave normalizada)
        key = normalize_key(titulo)
        raw_b64 = images.get(key)
        if raw_b64:
            try:
                if "," in raw_b64:
                    raw_b64 = raw_b64.split(",", 1)[1]
                img_data = base64.b64decode(raw_b64)
                img = ImageReader(io.BytesIO(img_data))
                img_w = right_limit - left_margin
                img_h = 200
                if y - img_h < 80:
                    c.showPage()
                    _new_page_with_background()
                    y = height - 80
                c.drawImage(img, left_margin, y - img_h, width=img_w, height=img_h, preserveAspectRatio=True, mask="auto")
                y -= img_h + 20

                # Si es el gráfico de pastel, añadir leyenda debajo de la imagen
                if "salon" in titulo.lower():
                    c.setFont("Helvetica-Bold", 9)
                    c.setFillColor(colors.HexColor("#cbd5e1"))
                    c.drawString(left_margin, y, "Referencia de colores:")
                    y -= 12
                    c.setFont("Helvetica", 8)
                    legend_y = y
                    for idx, fila in enumerate(datos):
                        salon_name = str(fila.get("Etiqueta", ""))
                        colors_list = [
                            "#6366F1", "#22C55E", "#EAB308", "#F97316", "#EF4444",
                            "#3B82F6", "#A855F7", "#10B981", "#F59E0B", "#8B5CF6",
                            "#EC4899", "#F43F5E", "#3B82F6", "#6366F1", "#22C55E",
                            "#F97316", "#EF4444", "#A855F7", "#10B981", "#F59E0B",
                            "#8B5CF6", "#EC4899", "#F43F5E", "#3B82F6", "#6366F1"
                        ]
                        color_hex = colors_list[idx % len(colors_list)]
                        c.setFillColor(colors.HexColor(color_hex))
                        c.rect(left_margin, legend_y - 6, 8, 8, fill=1, stroke=0)
                        c.setFillColor(colors.lightgrey)
                        c.drawString(left_margin + 12, legend_y - 4, salon_name)
                        legend_y -= 10
                        if legend_y < 100:
                            c.showPage()
                            _new_page_with_background()
                            legend_y = height - 80
                    y = legend_y - 10
            except Exception:
                c.setFillColor(colors.orange)
                c.setFont("Helvetica-Oblique", 9)
                c.drawString(left_margin, y, "(No se pudo insertar la imagen del chart)")
                y -= 16
                c.setFillColor(colors.lightgrey)
        else:
            raise ValueError(f"No se encontró imagen para el chart '{titulo}' (clave esperada: '{key}').")

    c.save()
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="reservaciones_periodo.pdf"'
    return response


# ===============================================================
# GENERADOR DE DATOS
# ===============================================================
def generar_datos_para(charts):
    base_data = [
        {"Etiqueta": "Lunes", "Valor": 12},
        {"Etiqueta": "Martes", "Valor": 18},
        {"Etiqueta": "Miércoles", "Valor": 9},
        {"Etiqueta": "Jueves", "Valor": 14},
        {"Etiqueta": "Viernes", "Valor": 11},
        {"Etiqueta": "Sábado", "Valor": 16},
        {"Etiqueta": "Domingo", "Valor": 8},
    ]

    charts_data = []
    for name in charts:
        charts_data.append({
            "nombre": name.replace("_", " ").capitalize(),
            "datos": base_data
        })
    return charts_data


# ===============================================================
# VISTAS
# ===============================================================
def descargar_csv(request):
    charts = request.GET.get("charts", "").split(",")
    return exportar_csv(generar_datos_para(charts))


def descargar_xlsx(request):
    charts = request.GET.get("charts", "").split(",")
    return exportar_xlsx_nativo(generar_datos_para(charts))

def descargar_reservaciones_periodo(request):
    return exportar_xlsx_reservaciones_semanal()



def descargar_pdf(request):
    """
    Recolección robusta de imágenes para generar PDF.
    Acepta:
      - POST form (multipart/form-data) con keys: bar, pie, line
      - POST form con keys: reservaciones, salones_mas_reservados, ingresos
      - POST field "images" con JSON {"bar": "...", ...} o {"reservaciones": "...", ...}
      - POST body JSON {"images": {...}}
      - GET ?images=JSON (fallback)
    Normaliza claves y devuelve 400 con mensaje claro si faltan imágenes.
    """
    import json
    import base64

    # DEBUG: mostrar claves recibidas en POST/FILES para detectar qué envía el frontend
    print("descargar_pdf - request.POST keys:", list(request.POST.keys()))
    print("descargar_pdf - request.FILES keys:", list(request.FILES.keys()))
    
    # keys que el PDF espera
    expected_keys = ["reservaciones", "salones_mas_reservados", "ingresos"]

    # mapa de alias que el frontend puede enviar
    key_map = {
        "bar": "reservaciones",
        "pie": "salones_mas_reservados",
        "line": "ingresos",
        "reservaciones_por_periodo": "reservaciones",
        "uso_de_salones": "salones_mas_reservados",
        "reservaciones": "reservaciones",
        "salones_mas_reservados": "salones_mas_reservados",
        "ingresos": "ingresos",
    }

    images = {}

    # 1) Si POST y existe campo 'images' con JSON
    if request.method == "POST":
        images_field = request.POST.get("images")
        if images_field:
            try:
                raw_images = json.loads(images_field)
                if isinstance(raw_images, dict):
                    for k, v in raw_images.items():
                        target = key_map.get(k, k)
                        if v:
                            images[target] = v
            except Exception:
                return HttpResponse("El campo 'images' debe contener JSON válido.", status=400)

        # 2) Remapear campos individuales en form-data (texto base64 o dataURL)
        for alias, target in key_map.items():
            if target in images:
                continue
            val = request.POST.get(alias)
            if val:
                images[target] = val

        # 3) Soportar archivos subidos en form-data (image files)
        for alias, target in key_map.items():
            if target in images:
                continue
            f = request.FILES.get(alias)
            if f:
                try:
                    raw = f.read()
                    images[target] = base64.b64encode(raw).decode("ascii")
                except Exception:
                    pass

        # 4) Si aún vacío, intentar parsear body como JSON
        if not images and request.body:
            try:
                body_json = json.loads(request.body.decode("utf-8"))
                raw_images = body_json.get("images") or body_json
                if isinstance(raw_images, dict):
                    for k, v in raw_images.items():
                        target = key_map.get(k, k)
                        if v:
                            images[target] = v
            except Exception:
                # no hacer nada; se validará más abajo
                pass

    else:
        # GET fallback: ?images=<json>
        try:
            raw_images = json.loads(request.GET.get("images", "{}"))
            if isinstance(raw_images, dict):
                for k, v in raw_images.items():
                    target = key_map.get(k, k)
                    if v:
                        images[target] = v
        except Exception:
            return HttpResponse("El parámetro 'images' debe ser JSON válido.", status=400)

    # limpiar vacíos
    images = {k: v for k, v in images.items() if v}

    # DEBUG: log keys recibidas (ver en consola)
    print("descargar_pdf - images keys:", list(images.keys()))

    # verificar que tengamos todas las keys esperadas (puedes relajar esto si quieres)
    missing = [k for k in expected_keys if k not in images]
    if missing:
        return HttpResponse(f"Faltan imágenes para: {', '.join(missing)}. Envíe 'bar','pie','line' o 'reservaciones','salones_mas_reservados','ingresos' o un campo JSON 'images'.", status=400)

    # charts: permitir que vengan en POST/GET, si no vienen usar defaults
    charts_param = request.POST.get("charts") if request.method == "POST" else request.GET.get("charts", "")
    charts = charts_param.split(",") if charts_param else []
    if not charts:
        charts = ["Reservaciones", "Salones mas reservados", "Ingresos"]

    fecha = request.POST.get("fecha") or request.GET.get("fecha")

    try:
        return exportar_pdf(generar_datos_para(charts), images, fecha_descarga=fecha)
    except ValueError as e:
        return HttpResponse(str(e), status=400)
    except Exception as e:
        # log del error para debugging
        import traceback; traceback.print_exc()
        return HttpResponse("Error al generar PDF", status=500)
